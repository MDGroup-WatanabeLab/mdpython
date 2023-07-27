from codecs import charmap_build
import os
import re
import openpyxl as px
from openpyxl.chart import BarChart, Series, Reference
from datetime import datetime, timedelta

file = "data.xlsx"
structure = []
step = []
user_time = []
config_num = []

date_old = []
run_num_old = []
structure_old = []
step_old = []
user_time_old = []
config_num_old = []

# for date in excel
def excel_date(date):
    if type(date) is int:
        tmp = datetime(1899, 12, 30) + timedelta(days=date)
        date = str(tmp.year) + "/" + str(tmp.month) + "/"  + str(tmp.day)
    return date

# check str or int
def check_int(text):
    if text.isdigit():
        return int(text)
    else:
        return text

# for sort list of directory
def natural_keys(dir_list):
    tmp = re.split(r'(\d+)', dir_list)
    result = []
    for i in tmp:
        result.append(check_int(i))
    return result

# get directory
path = os.getcwd()
dirlist = os.listdir(path)
dir_list = []
for i in dirlist:
    if os.path.isdir(i):
        dir_list.append(i)

# sort directory
dir_list = sorted(dir_list, key=natural_keys)

# get data
for dir in dir_list:
    if os.path.isfile(dir+"/INCAR"):
        f = open(dir+"/INCAR", "r")
        line = f.readlines()
        for i in line:
            if re.search("NSW", i):
                tmp = i.split()
                step.append(int(tmp[2]))
        f.close()
    else:
        step.append(-1)

    if os.path.isfile(dir+"/OUTCAR"):
        f = open(dir+"/OUTCAR", "r")
        line = f.readlines()
        for i in line:
            if re.search("User", i):
                tmp = i.split()
                user_time.append(float(tmp[3]))
        f.close()
    else:
        user_time.append(-1.0)

    if os.path.isfile(dir+"/ML_ABN"):
        f = open(dir+"/ML_ABN", "r")
        line = f.readlines()
        for i in range(len(line)):
            if re.search("The atom types in the data file", line[i]):
                structure.append(line[i+2].split())
            if re.search("The number of configuration", line[i]):
                config_num.append(int(line[i+2]))
        f.close()
    else:
        structure.append("Nofile")
        config_num.append(0)

# write excel file
if os.path.isfile(file):
    wb = px.load_workbook(file)
    ws = wb.worksheets[0]
    row_old = ws.max_row
    row = ws.max_row + 1

    # load excel file
    for i in range(row_old-1):
        if type(ws.cell(2+i, 2).value) is int:
            tmp = excel_date(ws.cell(2+i, 2).value)
        else:
            tmp = ws.cell(2+i, 2).value
        date_old.append(tmp)
        run_num_old.append(ws.cell(2+i, 3).value)
        structure_old.append(ws.cell(2+i, 4).value)
        step_old.append(ws.cell(2+i, 5).value)
        user_time_old.append(ws.cell(2+i, 6).value)
        config_num_old.append(ws.cell(2+i, 7).value)
    
    # create new sheet and delete old sheet
    # write data to new sheet
    wb.create_sheet("tmp")
    wb.remove(ws)
    ws = wb.worksheets[0]
    ws.title = "calc process"
    for i in range(row_old-1):
        ws.cell(2+i, 2).value = date_old[i]
        ws.cell(2+i, 3).value = run_num_old[i]
        ws.cell(2+i, 4).value = structure_old[i]
        ws.cell(2+i, 5).value = step_old[i]
        ws.cell(2+i, 6).value = user_time_old[i]
        ws.cell(2+i, 7).value = config_num_old[i]
    num = step_old[len(step_old)-1]

else:
    wb = px.Workbook()
    ws = wb.worksheets[0]
    ws.title = "calc process"
    ws.cell(2, 2).value = "date"
    ws.cell(2, 3).value = "number of run"
    ws.cell(2, 4).value = "structure"
    ws.cell(2, 5).value = "step"
    ws.cell(2, 6).value = "User time (sec)"
    ws.cell(2, 7).value = "number of structure"
    row = 3
    num = 0

# write new data
for i in range(len(step)):
    ws = wb.worksheets[0]
    tmp = ""
    ws.cell(row+i, 3).value = i+1+row-3
        
    for j in range(len(structure[i])):
        tmp = tmp + structure[i][j] + " "
    ws.cell(row+i, 4).value = tmp
    
    num += step[i]
    ws.cell(row+i, 5).value = num
        
    ws.cell(row+i, 6).value = user_time[i]
        
    ws.cell(row+i, 7).value = config_num[i]

# make bar chart
chart = BarChart()
chart.height = 20
chart.width = 20
m_row = ws.max_row
category = Reference(ws, min_col = 5, min_row = 3, max_row = m_row)
data = Reference(ws, min_col = 7, min_row = 3, max_row = m_row)
chart.add_data(data)
chart.set_categories(category)
chart.x_axis.title = 'The number of step'
chart.y_axis.title = 'The number of structure' 
chart.legend = None 
ws.add_chart(chart, "J2")
wb.save(file)