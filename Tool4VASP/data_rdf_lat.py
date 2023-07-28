import os
import re
import openpyxl as px
import shutil


file = "data_rdf_lat.xlsx"
plist = [1, 2, 4, 10, 20, 40, 100] # pick up directory
pcount = 0 # pick up counter 
path_vasp = "/home/owner/VASP/vasp.6.4.0/bin/vasp_std" # for 4.5
mpi_num = "16"
dir_excel = []
rdf_val = [] # store rdf values
LA = [] # lattice A
LB = [] # lattice B
LC = [] # lattice C
ALPHA = [] # angle B-C
BETA = [] # angle C-A
GAMMA = [] # angle A-B
LV = []   # lattice volume

# get directory name list
path = os.getcwd()
dirlist = os.listdir(path)
dir_list = []
for i in dirlist:
    if os.path.isdir(i):
        dir_list.append(i)

# check str or int
def check_int(text):
    if text.isdigit():
        return int(text)
    else:
        return text

# for sort list of directory
def natural_keys(dir_list):
    tmp = re.split(r'(\d+)', dir_list)
    result = []
    for i in tmp:
        result.append(check_int(i))
    return result

# sort directory
dir_list = sorted(dir_list, key=natural_keys)

#main loop for search for pick up the data
for dir in dir_list:
    if int(re.findall("[0-9]+$", dir)[0]) in plist: # decide pick up or not from plist
        dir_excel.append(dir)
        os.chdir(dir) # enter for directory
        dir_refit = dir + "_refit"
        os.mkdir(dir_refit)
        # copy 6 files
        shutil.copy("ICONST", dir_refit)
        shutil.copy("../INCAR", dir_refit)
        shutil.copy("../POSCAR", dir_refit) # use initial structure
        shutil.copy("POTCAR", dir_refit)
        shutil.copy("KPOINTS", dir_refit)
        shutil.copy("ML_ABN", dir_refit + "/ML_AB")

        # enter for directory
        os.chdir(dir_refit)
        # change to ML_MODE = REFIT in INCAR
        with open("INCAR", "r") as f:
            line = f.readlines()
            for i in range(len(line)):       
                if re.search("ML_MODE", line[i]):
                    line[i] = "ML_MODE = REFIT\n"
                if re.search("NSW", line[i]):
                    nsw_f = line[i].split()
                    nsw = str(nsw_f[2])
        with open("INCAR", "w") as f:
            for i in line:
                print(i, end = "", file = f)
            
        # start calculation for making ML_FFN
        os.system("mpirun -np " + mpi_num + " " + path_vasp)
        # finished calculation
        os.chdir("..")
        # back to main directory
        # make directory for ML_MODE = RUN
        dir_run = dir + "_run"
        os.mkdir(dir_run)
        # copy 7 files
        shutil.copy("ICONST", dir_run)
        shutil.copy("../INCAR", dir_run)
        shutil.copy("../POSCAR", dir_run) # use initial structure
        shutil.copy("POTCAR", dir_run)
        shutil.copy("KPOINTS", dir_run)
        #shutil.copy(dir_refit + "/ML_FFN", dir_run + "/ML_FF")
        shutil.copy("ML_FFN", dir_run + "/ML_FF")
        shutil.copy("../pair_correlation_xny.sh", dir_run)
        os.chdir(dir_run)
        # change to ML_MODE = RUN in INCAR
        with open("INCAR", "r") as f:
            line = f.readlines()
            for i in range(len(line)):       
                if re.search("ML_MODE", line[i]):
                    line[i]= "ML_MODE = RUN\n"
        with open("INCAR", "w") as f:
            for i in line:
                print(i, end = "", file = f)

        # start calculation for making ML_FFN
        os.system("mpirun -np " + mpi_num + " " + path_vasp)
        # finished calculation
        # make PCDAT.xy from PCDAT file
        os.system("bash pair_correlation_xny.sh")
        # store values
        with open("PCDAT.xy", "r") as f:
            line_rdf = f.readlines()
            for i in line_rdf:
                if re.search("final", i):
                    break
                rdf_val.append(i.split())
        
        with open("REPORT", "r") as f:
            line_lat = f.readlines()
            for i in range(len(line_lat)):
                if re.search("MD step No.", line_lat[i]):
                    if line_lat[i].split()[3] == nsw: # find final report block
                        la = line_lat[i+4].split()
                        LA.append(la[2])
                        lb = line_lat[i+5].split()
                        LB.append(lb[2])
                        lc = line_lat[i+6].split()
                        LC.append(lc[2])
                        alpha = line_lat[i+7].split()
                        ALPHA.append(alpha[2])
                        beta = line_lat[i+8].split()
                        BETA.append(beta[2])
                        gamma = line_lat[i+9].split()
                        GAMMA.append(gamma[2])
                        lv = line_lat[i+10].split()
                        LV.append(lv[2])
        os.chdir("../..") # back to main directory

        
print(rdf_val)
print(LA)
print(LB)
print(LC)
print(ALPHA)
print(BETA)
print(GAMMA)
print(LV)
print(dir_excel)

# make create Excel sheet
wb = px.Workbook()
ws_r = wb.worksheets[0]
ws_r.title = "rdf"
ws_r.cell(1, 1).value = "r [A]"
ws_r.cell(1, 2).value = "DFT"
ws_l = wb.create_sheet("lattice")

# rdf sheet
for i in range(len(rdf_val)):
    ws_r.cell((i%256)+2, 1).value = float(rdf_val[i][0]) # r [A]
    if i == 256:
        break

for i in range(len(rdf_val)):
    ws_r.cell(1, int(i/256)+1+len(rdf_val[i])).value = dir_excel[int(i/256)]
    for j in range(len(rdf_val[i])):
        if j != 0:
            # make space for DFT dataset
            ws_r.cell((i%256)+2, int(i/256)+1+len(rdf_val[i])).value = float(rdf_val[i][j])

# lattice sheet
ws_l.cell(2, 1).value = "LA"
ws_l.cell(3, 1).value = "LB"
ws_l.cell(4, 1).value = "LC"
ws_l.cell(5, 1).value = "alpha"
ws_l.cell(6, 1).value = "beta"
ws_l.cell(7, 1).value = "gamma"
ws_l.cell(8, 1).value = "LV"
ws_l.cell(1, 2).value = "DFT"

for i in range(len(dir_excel)):
    ws_l.cell(1, 3+i).value = dir_excel[i]
    ws_l.cell(2, 3+i).value = float(LA[i])
    ws_l.cell(3, 3+i).value = float(LB[i])
    ws_l.cell(4, 3+i).value = float(LC[i])
    ws_l.cell(5, 3+i).value = float(ALPHA[i])
    ws_l.cell(6, 3+i).value = float(BETA[i])
    ws_l.cell(7, 3+i).value = float(GAMMA[i])
    ws_l.cell(8, 3+i).value = float(LV[i])

wb.save(file)



                


            






        


