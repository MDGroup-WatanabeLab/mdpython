from codecs import charmap_build
import os
import re
import openpyxl as px
from openpyxl.chart import BarChart, Series, Reference
from datetime import datetime, timedelta

file = "data_err.xlsx"
err_list = []
beef_list = []
ctifor_list = []
step_list_1 = [] # for ERR
step_list_2 = [] # for BEEF



# check str or int
def check_int(text):
    if text.isdigit():
        return int(text)
    else:
        return text

# for sort list of directory
def natural_keys(dir_list):
    tmp = re.split(r'(\d+)', dir_list)
    result = []
    for i in tmp:
        result.append(check_int(i))
    return result

# get directory
path = os.getcwd()
dirlist = os.listdir(path)
dir_list = []
for i in dirlist:
    if os.path.isdir(i):
        dir_list.append(i)

# sort directory
dir_list = sorted(dir_list, key=natural_keys)

# get data
for dir in dir_list:
    # get timestep 
    if os.path.isfile(dir+"/INCAR"):
        with open(dir+"/INCAR", "r") as f:
            line = f.readlines()
            for i in line:
                if re.search("NSW", i):
                    tmp = i.split()
                    step = int(tmp[2])
    #grep ERR ML_LOGFILE|grep -v "#"|awk '{print $2, $4}' > ERR.dat
    #grep BEEF ML_LOGFILE|grep -v "#"|awk '{print $2, $4}' > BEEF.dat
    #grep BEEF ML_LOGFILE|grep -v "#"|awk '{print $2, $6}' > CTIFOR.dat

    if os.path.isfile(dir+"/ML_LOGFILE"):
        step_num = re.split(r'(\d+)', dir)
        with open(dir+"/ML_LOGFILE", "r") as f:
            line = f.readlines()
            for i in line:
                # get ERR
                if re.search("ERR", i):
                    tempo = i.split()
                    if tempo[0] == "ERR":
                        step_list_1.append(int(tempo[1]) + (int(step_num[1])-1)*step)
                        err_list.append(tempo[3])
                # get BEEF
                # get CTIFOR
                if re.search("BEEF", i):
                    tempo2 = i.split()
                    if tempo2[0] == "BEEF":
                        step_list_2.append(int(tempo2[1]) + (int(step_num[1])-1)*step)
                        beef_list.append(tempo2[3])
                        ctifor_list.append(tempo2[5])


# write excel file
wb = px.Workbook()
# write new data
row = 2
# for ERR
ws_err = wb.worksheets[0]
ws_err.title = "ERROR"
ws_err.cell(1, 1).value = "number of run"
ws_err.cell(1, 2).value = "ERR"
for i in range(len(step_list_1)):
    ws_err.cell(row+i, 1).value = int(step_list_1[i])
    ws_err.cell(row+i, 2).value = float(err_list[i])

# for BEEF and CTIFOR
ws_beef = wb.create_sheet("BEEF")
ws_beef.cell(1, 1).value = "number of run"
ws_beef.cell(1, 2).value = "BEEF"
ws_beef.cell(1, 3).value = "CTIFOR"
for i in range(len(step_list_2)):
    ws_beef.cell(row+i, 1).value = int(step_list_2[i])
    ws_beef.cell(row+i, 2).value = float(beef_list[i])
    ws_beef.cell(row+i, 3).value = float(ctifor_list[i])

# for all 3 values
err_list_new = [list(e) for e in zip(step_list_1, err_list)]
print(err_list_new)

# for BEEF and CTIFOR
ws_all = wb.create_sheet("ALL")
ws_all.cell(1, 1).value = "number of run"
ws_all.cell(1, 2).value = "ERR"
ws_all.cell(1, 3).value = "BEEF"
ws_all.cell(1, 4).value = "CTIFOR"
for i in range(len(step_list_2)):
    ws_all.cell(row+i, 1).value = int(step_list_2[i])
    ws_all.cell(row+i, 3).value = float(beef_list[i])
    ws_all.cell(row+i, 4).value = float(ctifor_list[i])

for i in range(len(err_list_new)):
    ws_all.cell(row-1+int(err_list_new[i][0]), 2).value = float(err_list_new[i][1])

# Make scatter graph
#definition of graph properties
chart=px.chart.ScatterChart()
chart.title = "3 Errors in force"
chart.x_axis.title = "timestep"
chart.y_axis.title = "Error in force [eV/Angstrom]"
chart.legend.position = "b"
chart.smooth = True
chart.width = 24
chart.height = 16

#Select area for making graph
y_err = px.chart.Reference(wb["ALL"] ,min_col=2 ,max_col=2 ,min_row=2 ,max_row= len(step_list_1)+1)
y_beef = px.chart.Reference(wb["ALL"] ,min_col=3 ,max_col=3 ,min_row=2 ,max_row= len(step_list_1)+1)
y_ctifor = px.chart.Reference(wb["ALL"] ,min_col=4 ,max_col=4 ,min_row=2 ,max_row= len(step_list_1)+1)
x_err = px.chart.Reference(wb["ALL"] ,min_col=1 ,max_col=1 ,min_row=2 ,max_row= len(step_list_1)+1)

#系列変数seriesをy,xを指定して定義する
series_err = px.chart.Series(y_err, x_err, title="ERROR")
series_beef = px.chart.Series(y_beef, x_err, title="BEEF")
series_ctifor = px.chart.Series(y_ctifor, x_err, title="CTIFOR")

#Delete marker

series_err.marker.symbol = None
series_beef.marker.symbol = None
series_ctifor.marker.symbol = None
chart.series.append(series_err)
chart.series.append(series_beef)
chart.series.append(series_ctifor)

#Show graph at D3
wb["ALL"].add_chart(chart,"F1")


wb.save(file)

wb.close()