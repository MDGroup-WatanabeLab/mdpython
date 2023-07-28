import openpyxl as px
from openpyxl.chart import ScatterChart, LineChart, Reference, Series

f = open("EIGENVAL", "r")
line = f.read().splitlines()
length = len(line)

# 電子数、K点数、バンド数を取得
elec, kpointnum, bandnum = line[5].split()
elec =int(elec)
kpointnum = int(kpointnum)
bandnum = int(bandnum)

#必要な配列の宣言、エネルギー等は動的に確保する
kx = [0] * kpointnum
ky = [0] * kpointnum
kz = [0] * kpointnum
w = [0] * kpointnum

# factorは各バンドのエネルギーを格納する「energy」の要素を示すためのもの
for i in range(bandnum):
    exec("energy{} = [0] * kpointnum".format(str(i+1)))
    exec("state{} = [0] * kpointnum".format(str(i+1)))
    exec("factor{} = 0".format(str(i+1)))

num = 0
factor_k = 0

# K点の座標の取得、各バンドごとのエネルギーに分ける
for i in range(length):
    if i >= 7:
        if num % (bandnum + 2) == 0:
            kx[factor_k], ky[factor_k], kz[factor_k], w[factor_k] = line[i].split()
            kx[factor_k] = float(kx[factor_k])
            ky[factor_k] = float(ky[factor_k])
            kz[factor_k] = float(kz[factor_k])
            factor_k += 1
            num += 1
        elif num == bandnum + 1:
            num = 0
        else:
            tmp1, *tmp2 = line[i].split()
            exec("energy{}[factor{}] = tmp2[0]".format(tmp1, tmp1))
            exec("state{}[factor{}] = tmp2[1]".format(tmp1, tmp1))
            exec("factor{} += 1".format(tmp1))
            num += 1

# 型変換
for i in range(bandnum):
    exec("factor{} = 0".format(str(i+1)))
    for j in range(kpointnum):
        exec("energy{}[j] = float(energy{}[j])".format(str(i + 1), str(i + 1)))

f.close()

#エクセルへの出力とバンド構造の図示
wb = px.Workbook()
ws = wb.worksheets[0]
ws.title = ("EIGENVAL")

for i in range(bandnum):
    bandname = 'band' + str(i + 1)
    ws.cell(1, i + 2).value = bandname

ws.cell(1, bandnum + 3).value = "k_x"
ws.cell(1, bandnum + 4).value = "k_y"
ws.cell(1, bandnum + 5).value = "k_z"

for i in range(kpointnum):
    ws.cell(i + 2, 1).value = i
    ws.cell(2 + i, bandnum + 3).value = kx[i]
    ws.cell(2 + i, bandnum + 4).value = ky[i]
    ws.cell(2 + i, bandnum + 5).value = kz[i]
    for j in range(bandnum):
        exec("ws.cell({}, {}).value = energy{}[factor{}]".format(i + 2, j + 2, str(j + 1), str(j + 1)))
        exec("factor{} += 1".format(str(j + 1)))

chart = ScatterChart()
chart.height = 20
chart.width = 30
min_row = 2
max_row = min_row + kpointnum - 1
x_values = Reference(ws, min_col = 1, min_row = min_row, max_row = max_row)

for i in range(bandnum):
    y_values = Reference(ws, min_col = 2 + i, min_row = 1, max_row = max_row)
    series = Series(y_values, x_values, title_from_data = True)
    chart.series.append(series)

chart.x_axis.title = ''
chart.y_axis.title = 'Energy'
chart.x_axis.crosses = "min"
chart.y_axis.crosses = "min"
chart.x_axis.scaling.max = kpointnum
ws.add_chart(chart, "D5")

wb.save("EIGENVAL.xlsx")

# reference
# https://www.youtube.com/watch?v=-V-bwX8tpkk