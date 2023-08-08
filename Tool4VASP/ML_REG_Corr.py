import pandas as pd
import openpyxl as px
from openpyxl.chart import ScatterChart, LineChart, Reference, Series

f = open('ML_REG', 'r')
list = f.read().splitlines()

energy = []
force = []
stress = []
length_list = len(list)

# エネルギー、力、応力を配列に格納
flag = 0
i = 0
for j in range(length_list):
    if j > 4:
        check = list[j][:1]
        if check == "=":
            i += 3
            flag += 1
        elif i == j:
            if flag == 0:
                energy.append((list[j].rstrip("\n")))
            elif flag == 1:
                force.append((list[j].rstrip("\n")))
            elif flag == 2:
                stress.append((list[j].rstrip("\n")))
    if i == j:
        i += 1

# 一列目と二列目を分離
# 浮動小数に変換
length_energy = len(energy)
x_energy = [0] * length_energy
y_energy = [0] * length_energy
for a in range(length_energy):
    x_energy[a], y_energy[a] = energy[a].split()
    x_energy[a] = float(x_energy[a])
    y_energy[a] = float(y_energy[a])

length_force = len(force)
x_force = [0] * length_force
y_force = [0] * length_force
for b in range(length_force):
    x_force[b], y_force[b] = force[b].split()
    x_force[b] = float(x_force[b])
    y_force[b] = float(y_force[b])
    
length_stress = len(stress)
x_stress = [0] * length_stress
y_stress = [0] * length_stress
for c in range(length_stress):
    x_stress[c], y_stress[c] = stress[c].split()
    x_stress[c] = float(x_stress[c])
    y_stress[c] = float(y_stress[c])

wb = px.Workbook()

# エクセルにデータを書き込み、グラフを作成
for type in range(6):
    row = 1
    column = 1
    if type == 0:
        ws = wb.worksheets[0]
        ws.title = "Energy"
        ac = ws.cell(row = row, column = column)
        ac.value = "Validate"
        
        for num in x_energy:
            row += 1
            ac = ws.cell(row = row, column = column)      
            ac.value = num
            
    elif type == 1:
        column += 1
        ac = ws.cell(row = row, column = column)
        ac.value = "Training"
        
        for num in y_energy:
            row += 1
            ac = ws.cell(row = row, column = column)
            ac.value = num
        
        chart = ScatterChart()
        chart.height = 20
        chart.width = 20
        min_row = 2
        max_row = min_row + length_energy - 1
        x_values = Reference(ws, min_col = 1, min_row = min_row, max_row = max_row)
        y_values = Reference(ws, min_col = 2, min_row = min_row, max_row = max_row)
        series = Series(y_values, x_values)
        chart.series.append(series)
        chart.x_axis.title = 'Validate'
        chart.y_axis.title = 'Training'
        chart.x_axis.crosses = "min"
        chart.y_axis.crosses = "min"  
        chart.legend = None 
        for val in chart.series:
            val.marker.symbol = 'circle'
            val.graphicalProperties.line.noFill = True    
        ws.add_chart(chart, "D2")
        
    elif type == 2:
        ws = wb.create_sheet(title="Force")
        ac = ws.cell(row = row, column = column)
        ac.value = "Validate"
        
        for num in x_force:
            row += 1
            ac = ws.cell(row = row, column = column)
            ac.value = num
            
    elif type == 3:
        column += 1
        ac = ws.cell(row = row, column = column)
        ac.value = "Training"
        
        for num in y_force:
            row += 1
            ac = ws.cell(row = row, column = column)
            ac.value = num
            
        chart = ScatterChart()
        chart.height = 20
        chart.width = 20
        min_row = 2
        max_row = min_row + length_force - 1
        x_values = Reference(ws, min_col = 1, min_row = min_row, max_row = max_row)
        y_values = Reference(ws, min_col = 2, min_row = min_row, max_row = max_row)
        series = Series(y_values, x_values)
        chart.series.append(series)
        chart.x_axis.title = 'Validate'
        chart.y_axis.title = 'Training'
        chart.x_axis.crosses = "min"
        chart.y_axis.crosses = "min"  
        chart.legend = None 
        for val in chart.series:
            val.marker.symbol = 'circle'
            val.graphicalProperties.line.noFill = True    
        ws.add_chart(chart, "D2")
            
    elif type == 4:
        ws = wb.create_sheet(title="Stress")
        ac = ws.cell(row = row, column = column)
        ac.value = "Validate"
        
        for num in x_stress:
            row += 1
            ac = ws.cell(row = row, column = column)
            ac.value = num
            
    elif type == 5:
        column += 1
        ac = ws.cell(row = row, column = column)
        ac.value = "Training"
        
        for num in y_stress:
            row += 1
            ac = ws.cell(row = row, column = column)
            ac.value = num
            
        chart = ScatterChart()
        chart.height = 20
        chart.width = 20
        min_row = 2
        max_row = min_row + length_stress - 1
        x_values = Reference(ws, min_col = 1, min_row = min_row, max_row = max_row)
        y_values = Reference(ws, min_col = 2, min_row = min_row, max_row = max_row)
        series = Series(y_values, x_values)
        chart.series.append(series)
        chart.x_axis.title = 'Validate'
        chart.y_axis.title = 'Training'
        chart.x_axis.crosses = "min"
        chart.y_axis.crosses = "min"  
        chart.legend = None 
        for val in chart.series:
            val.marker.symbol = 'circle'
            val.graphicalProperties.line.noFill = True    
        ws.add_chart(chart, "D2")
            

# 相関係数をエクセルに書き込む
ws = wb.create_sheet(title="相関係数")
wb.save('ML_REG_Corr.xlsx')

df_enegy = pd.DataFrame({'Validate':x_energy, 'Training':y_energy})
df_corr_energy = df_enegy.corr()
df_force = pd.DataFrame({'Validate':x_force, 'Training':y_force})
df_corr_force = df_force.corr()
df_stress = pd.DataFrame({'Validate':x_stress, 'Training':y_stress})
df_corr_stress = df_stress.corr()

with pd.ExcelWriter("ML_REG_Corr.xlsx", mode = "a", if_sheet_exists="overlay", engine = "openpyxl") as writer:
    df_corr_energy.to_excel(writer, sheet_name = "相関係数", index_label = "Energy")
    df_corr_force.to_excel(writer, sheet_name = "相関係数", startrow = 4, index_label = "Force")
    df_corr_stress.to_excel(writer, sheet_name = "相関係数", startrow = 8, index_label = "Stress")
