#general import
import re
import openpyxl as px
import os

# make new workbook
wb = px.Workbook()

#################### Presure Block ####################
if os.path.exists("OUTCAR") == True:
    f_p = open("OUTCAR", "r")
    list_p = f_p.readlines()
    num_p = []
    press = []

    # get total pressure value 
    for i in list_p:
        if re.search("total pressure", i):
            *_, press, _ = i.split()              #In _ , "total presssure =" and "kB"
            num_p.append(press)



    # Write worksheet
    ws_p = wb.worksheets[0]
    ws_p.title = "Pressure (kB)"
    ws_p.cell(1, 1).value = "Timestep"
    ws_p.cell(1, 2).value = "Pressure"

    for i in range(len(num_p)):
        ws_p.cell(i+2, 1).value = i + 1
        ws_p.cell(i+2, 2).value = float(num_p[i])

    # Make scatter graph
    #definition of graph properties
    chart_p=px.chart.ScatterChart()
    chart_p.title = "Pressure Change"
    chart_p.x_axis.title = "timestep"
    chart_p.y_axis.title = "pressure (kB)"
    chart_p.legend = None
    chart_p.smooth = True
    chart_p.width = 12
    chart_p.height = 8

    #Select area for making graph
    y_t = px.chart.Reference(wb["Pressure (kB)"] ,min_col=2 ,max_col=2 ,min_row=2 ,max_row= len(num_p)+1)
    x_t = px.chart.Reference(wb["Pressure (kB)"] ,min_col=1 ,max_col=1 ,min_row=2 ,max_row= len(num_p)+1)

    #系列変数seriesをy,xを指定して定義する
    series_p = px.chart.Series(y_t, x_t)

    #Delete marker
    series_p.marker.symbol = None
    chart_p.series.append(series_p)

    #Show graph at D3
    wb["Pressure (kB)"].add_chart(chart_p,"D3")

    wb.save("PVTE_graph.xlsx")

    f_p.close()



#################### Temperature Block ####################
if os.path.exists("OSZICAR") == True:
    f_t = open("OSZICAR", "r")

    list_t = f_t.readlines()
    num_t = []
    temp = []

    # get temperature value 
    for i in list_t:
        if re.search("T=", i):
            _, _, temp, *_ = i.split()
            num_t.append(temp)


    # Write worksheet
    ws_t = wb.create_sheet("Temperature")
    ws_t.title = "Temperature (K)"
    ws_t.cell(1, 1).value = "Timestep"
    ws_t.cell(1, 2).value = "Temperature"
    for i in range(len(num_t)):
        ws_t.cell(i+2, 1).value = i + 1
        ws_t.cell(i+2, 2).value = float(num_t[i])

    # Make scatter graph
    #definition of graph properties
    chart_t=px.chart.ScatterChart()
    chart_t.title = "Temperature Change"
    chart_t.x_axis.title = "timestep"
    chart_t.y_axis.title = "temperature (K)"
    chart_t.legend = None
    chart_t.smooth = True
    chart_t.width = 12
    chart_t.height = 8

    #Select area for making graph
    y_t = px.chart.Reference(wb["Temperature (K)"] ,min_col=2 ,max_col=2 ,min_row=2 ,max_row= len(num_t)+1)
    x_t = px.chart.Reference(wb["Temperature (K)"] ,min_col=1 ,max_col=1 ,min_row=2 ,max_row= len(num_t)+1)

    #系列変数seriesをy,xを指定して定義する
    series_t = px.chart.Series(y_t, x_t)

    #Delete marker
    series_t.marker.symbol = None
    chart_t.series.append(series_t)

    #Show graph at D3
    wb["Temperature (K)"].add_chart(chart_t,"D3")


    wb.save("PVTE_graph.xlsx")

    f_t.close()


#################### Volume Block ####################
if os.path.exists("REPORT") == True:
    f_v = open("REPORT", "r")

    list_v = f_v.readlines()
    num_v = []
    volume = []

    # get volume  value 
    for i in list_v:
        if re.search("LV", i):
            *_, volume = i.split()
            num_v.append(volume)


    # Write worksheet
    ws_v = wb.create_sheet("Volume")
    ws_v.title = "Volume (Ang.^3)"
    ws_v.cell(1, 1).value = "Timestep"
    ws_v.cell(1, 2).value = "Volume"
    for i in range(len(num_v)):
        ws_v.cell(i+2, 1).value = i + 1
        ws_v.cell(i+2, 2).value = float(num_v[i])

    # Make scatter graph
    #definition of graph properties
    chart_v=px.chart.ScatterChart()
    chart_v.title = "Volume Change"
    chart_v.x_axis.title = "timestep"
    chart_v.y_axis.title = "volume (Ang.^3)"
    chart_v.legend = None
    chart_v.smooth = True
    chart_v.width = 12
    chart_v.height = 8

    #Select area for making graph
    y_v = px.chart.Reference(wb["Volume (Ang.^3)"] ,min_col=2 ,max_col=2 ,min_row=2 ,max_row= len(num_v)+1)
    x_v = px.chart.Reference(wb["Volume (Ang.^3)"] ,min_col=1 ,max_col=1 ,min_row=2 ,max_row= len(num_v)+1)

    #系列変数seriesをy,xを指定して定義する
    series_v = px.chart.Series(y_v, x_v)

    #Delete marker
    series_v.marker.symbol = None
    chart_v.series.append(series_v)

    #Show graph at D3
    wb["Volume (Ang.^3)"].add_chart(chart_v,"D3")

    wb.save("PVTE_graph.xlsx")

    f_v.close()

#################### Energy Block ####################
if os.path.exists("REPORT") == True:
    f_et = open("REPORT", "r")
    f_ep = open("REPORT", "r")
    f_ek = open("REPORT", "r")

    list_et = f_et.readlines()
    list_ep = f_ep.readlines()
    list_ek = f_ek.readlines()
    num_et = []
    num_ep = []
    num_ek = []
    energy_t = []
    energy_p = []
    energy_k = []

    # get total energy value 
    for i in list_et:
        if re.search("e_b>", i):
            _, energy_t, *_ = i.split()
            num_et.append(energy_t)

    # get potential energy value 
    for i in list_ep:
        if re.search("e_b>", i):
            _, _, energy_p, _, _ = i.split()
            num_ep.append(energy_p)

    # get kinetic energy value 
    for i in list_ek:
        if re.search("e_b>", i):
            _, _, _, energy_k, _ = i.split()
            num_ek.append(energy_k)


    # Write worksheet
    ws_e = wb.create_sheet("Energy")
    ws_e.title = "Energy (eV)"
    ws_e.cell(1, 1).value = "Timestep"
    ws_e.cell(1, 2).value = "Total energy"
    ws_e.cell(1, 3).value = "Potential energy"
    ws_e.cell(1, 4).value = "Kinetic energy"

    for i in range(len(num_et)):
        ws_e.cell(i+2, 1).value = i + 1
        ws_e.cell(i+2, 2).value = float(num_et[i])

    for i in range(len(num_ep)):
        ws_e.cell(i+2, 1).value = i + 1
        ws_e.cell(i+2, 3).value = float(num_ep[i])

    for i in range(len(num_ek)):
        ws_e.cell(i+2, 1).value = i + 1
        ws_e.cell(i+2, 4).value = float(num_ek[i])

    # Make scatter graph
    #definition of graph properties
    chart_e=px.chart.ScatterChart()
    chart_e.title = "Energy Change"
    chart_e.x_axis.title = "timestep"
    chart_e.y_axis.title = "energy (eV)"
    chart_e.smooth = True
    chart_e.legend.position = "b"
    chart_e.width = 12
    chart_e.height = 8

    #Select area for making graph
    y_et = px.chart.Reference(wb["Energy (eV)"] ,min_col=2 ,max_col=2 ,min_row=2 ,max_row= len(num_et)+1)
    y_ep = px.chart.Reference(wb["Energy (eV)"] ,min_col=3 ,max_col=3 ,min_row=2 ,max_row= len(num_ep)+1)
    y_ek = px.chart.Reference(wb["Energy (eV)"] ,min_col=4 ,max_col=4 ,min_row=2 ,max_row= len(num_ek)+1)
    x_e = px.chart.Reference(wb["Energy (eV)"] ,min_col=1 ,max_col=1 ,min_row=2 ,max_row= len(num_et)+1)

    #系列変数seriesをy,xを指定して定義する
    series_et = px.chart.Series(y_et, x_e, title="Total energy")
    series_ep = px.chart.Series(y_ep, x_e, title="Potential energy")
    series_ek = px.chart.Series(y_ek, x_e, title="Kinetic energy")

    #Delete marker
    series_et.marker.symbol = None
    series_ep.marker.symbol = None
    series_ek.marker.symbol = None
    chart_e.series.append(series_et)
    chart_e.series.append(series_ep)
    chart_e.series.append(series_ek)

    #Show graph 
    wb["Energy (eV)"].add_chart(chart_e,"F3")

    wb.save("PVTE_graph.xlsx")

    wb.close()

    f_et.close()
    f_ep.close()
    f_ek.close()
