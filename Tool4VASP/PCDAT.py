import math
import openpyxl as px
from openpyxl.chart import ScatterChart, LineChart, Reference, Series

f = open("PCDAT", "r")
line = f.read().splitlines()
length = len(line)

check = line[12].split()
len_check = len(check)

# pairnumは原子の種類
# npacoはpcdatの行数、pcskalはfixed output(多分単位をÅするため？)
# pcskalはペア相関関数の横軸の間隔(pcskal = APACO * 10^-10 / npaco, APACO : ペア相関関数の最大距離(横軸をどこまで取るか))
type = (-1 + math.sqrt(1 + 8 * (len_check - 1))) / 2
npaco = int(line[6])
pcskal = float(line[7])
pcfein = float(line[8])

total = [0] * npaco
pair = [0] * npaco
r = [0] * npaco

for i in range(int(type)):
    for j in range(int(type)):
        if j >= i:
            exec("pair{}_{} = [0] * npaco".format(str(i + 1), str(j + 1)))

# 値の読み込み
j = 0
for i in range(length):
    if i >= 12:
        if int(len_check) == 1:
            total[j] = line[i]
            j += 1
        else:
            total[j], *pair[j] = line[i].split()
            j += 1

for i in range(npaco):
    r[i] = ((i + 1) - 0.5) * pcfein / pcskal
    total[i] = float(total[i])
    tmp = 0
    for j in range(int(type)):
        for k in range(int(type)):
            if k >= j:
                exec("pair{}_{}[i] = pair[i][tmp]".format(str(j + 1), str(k + 1)))
                exec("pair{}_{}[i] = float(pair{}_{}[i])".format(str(j + 1), str(k + 1), str(j + 1), str(k + 1)))
                tmp += 1

f.close()

# エクセルへの書き込み
wb = px.Workbook()
ws = wb.worksheets[0]
ws.title = "Pair Correlation Function"

ws.cell(1, 1).value = "Radius"
ws.cell(1, 2).value = "Total"
tmp = 0
for i in range(int(type)):
    for j in range(int(type)):
        if j >= i:
            exec("ws.cell(1, 3 + tmp).value = \"{}-{}\"".format(str(i + 1), str(j + 1)))
            tmp += 1

if int(len_check) == 1:
    for i in range(npaco):
        ws.cell(2 + i, 1).value = r[i]
        ws.cell(2 + i, 2).value = total[i]
else:
    for i in range(npaco):
        ws.cell(2 + i, 1).value = r[i]
        ws.cell(2 + i, 2).value = total[i]
        tmp = 0
        for j in range(int(type)):
            for k in range(int(type)):
                if k >= j:
                    exec("ws.cell(2 + i, 3 + tmp).value = pair{}_{}[i]".format(str(j + 1), str(k + 1)))
                    tmp += 1

chart = ScatterChart()
chart.height = 20
chart.width = 30
min_row = 2
max_row = min_row + npaco - 1
x_values = Reference(ws, min_col = 1, min_row = min_row, max_row = max_row)
y_values = Reference(ws, min_col = 2, min_row = 1, max_row = max_row)
series = Series(y_values, x_values, title_from_data = True)
chart.series.append(series)

for i in range(int(type * (type + 1) / 2)):
    y_values = Reference(ws, min_col = 3 + i, min_row = 1, max_row = max_row)
    series = Series(y_values, x_values, title_from_data = True)
    chart.series.append(series)

chart.x_axis.title = 'Radius'
chart.y_axis.title = 'Pair Correlation Function'
chart.x_axis.crosses = "min"
chart.y_axis.crosses = "min"
chart.title = None
ws.add_chart(chart, "D5")

wb.save("PCDAT.xlsx")