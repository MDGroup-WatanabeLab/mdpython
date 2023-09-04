import os
import re
import math
import openpyxl as px
from openpyxl.chart import ScatterChart, Reference, Series

# check str or int
def check_int(text):
    if text.isdigit():
        return int(text)
    else:
        return text

# for sort list of directory
def natural_keys(dir_list):
    tmp = re.split(r'(\d+)', dir_list)
    result = []
    for i in tmp:
        result.append(check_int(i))
    return result

# get directory
path = os.getcwd()
dirlist = os.listdir(path)
dir_list = []
for i in dirlist:
    if os.path.isdir(i):
        dir_list.append(i)
        
# sort directory
dir_list = sorted(dir_list, key=natural_keys)

volume = []
index_num = []
energy = []

# get data
for dir in dir_list:
    if os.path.isfile(dir+"/INCAR"):
        with open(dir+"/INCAR", "r") as f:
            line = f.readlines()
            for i in line:
                if re.search("NSW", i):
                    nsw = i.split()[2]
    
    if os.path.isfile(dir+"/REPORT") and os.path.getsize(dir+"/REPORT") != 0:
        tmp = []
        with open(dir+"/REPORT", "r") as f:
            line = f.readlines()
            for i in line:
                if re.search("mc> LV", i):
                    tmp.append(i.split()[2])
            volume.append(tmp[-1])
    elif os.path.isfile(dir+"/OUTCAR"):
        tmp = []
        with open(dir+"/OUTCAR", "r") as f:
            line = f.readlines()
            for i in line:
                if re.search("volume of cell", i):
                    tmp.append(i.split()[4])
            volume.append(tmp[-1])   
    
    if os.path.isfile(dir+"/OSZICAR"):
        tmp = []
        with open(dir+"/OSZICAR", "r") as f:
            line = f.readlines()
            for i in line:
                if re.search("F=", i):
                    tmp.append(i.split())
            index_num.append(tmp[-1].index("E0="))
            energy.append(tmp[-1][index_num[-1]+1])

wb = px.Workbook()
ws = wb.worksheets[0]
ws.title = "E-V_graph"
ws.cell(2, 2).value = "volume"
ws.cell(2, 3).value = "energy"

for i in range(len(volume)):
    ws.cell(3+i, 2).value = float(volume[i])
    ws.cell(3+i, 3).value = float(energy[i])

# make bar chart
file = "EV_graph.xlsx"
chart = ScatterChart()
chart.height = 20
chart.width = 20
m_row = ws.max_row
x = Reference(ws, min_col = 2, min_row = 3, max_row = m_row)
y = Reference(ws, min_col = 3, min_row = 3, max_row = m_row)
series = Series(y, x)
series.marker.symbol = "circle"
chart.series.append(series)
chart.x_axis.title = 'Volume'
chart.y_axis.title = 'Energy' 
chart.legend = None 

# dtermine min and max
v_min = math.floor(float(volume[0]))
v_max = math.ceil(float(volume[-1]))
v_min_len = len(str(v_min))
v_max_len = len(str(v_max))
if v_min_len > 2:
    v_min = math.floor(v_min/pow(10, v_min_len-2))*pow(10, v_min_len-2)
if v_max_len > 2:
    v_max = math.ceil(v_max/pow(10, v_max_len-2))*pow(10, v_max_len-2)

chart.x_axis.scaling.min = v_min
chart.x_axis.scaling.max = v_max
ws.add_chart(chart, "E2")
wb.save(file)
