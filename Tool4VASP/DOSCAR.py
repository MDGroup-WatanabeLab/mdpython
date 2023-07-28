import openpyxl as px
from openpyxl.chart import ScatterChart, LineChart, Reference, Series

f = open("DOSCAR", "r")
line = f.read().splitlines()
length = len(line)

*_, numdos, Fermi, potim = line[5].split()
numdos = int(numdos)
check = line[6].split()
len_check = len(check)

# エネルギーとDOS用の配列
energy = [0] * numdos
if len_check == 3:
    dos = [0] * numdos
    idos = [0] * numdos
else:
    dos_up = [0] * numdos
    dos_down = [0] * numdos
    idos_up = [0] * numdos
    idos_down = [0] * numdos

# DOSとDOSの積分以降は各原子種の軌道ごとのDOSが記述される？
# 構造を構成する原子の種類を取得する
type = (length - 5 - (numdos + 1)) / (numdos + 1)

if int(type) == 0:
    type1 = ["0"] * numdos
else:
    for i in range(int(type)):
        exec("type{} = [0] * numdos".format(str(i + 1)))

tmp = 0
flag = 0

# 値の読み込み
for i in range(length):
    if i >= 6:
        if tmp == numdos:
            tmp = 0
            flag += 1
        elif flag == 0:
            if len_check == 3:
                energy[tmp], dos[tmp], idos[tmp] = line[i].split()
                energy[tmp] = float(energy[tmp])
                dos[tmp] = float(dos[tmp])
                idos[tmp] = float(idos[tmp])
            else:
                energy[tmp], dos_up[tmp], dos_down[tmp], idos_up[tmp], idos_down[tmp] = line[i].split()
                energy[tmp] = float(energy[tmp])
                dos_up[tmp] = float(dos_up[tmp])
                dos_down[tmp] = float(dos_down[tmp])
                idos_up[tmp] = float(idos_up[tmp])
                idos_down[tmp] = float(idos_down[tmp])
            tmp += 1
        else:
            _, *tmp_line = line[i].split()
            exec("type{}[tmp] = tmp_line".format(str(flag)))
            tmp += 1

dos_type = len(type1[0])

f.close()

# エクセルへの出力と状態密度の図示
wb = px.Workbook()
ws = wb.worksheets[0]
ws.title = ("DOS")
ws.cell(1, 1).value = "Energy"
if len_check == 3:
    ws.cell(1, 2).value = "DOS"
    ws.cell(1, 3).value = "integrated DOS"
else:
    ws.cell(1, 2).value = "DOS(UP)"
    ws.cell(1, 3).value = "DOS(DOWN)"
    ws.cell(1, 4).value = "integrated DOS(UP)"
    ws.cell(1, 5).value = "integrated DOS(DOWN)"

for i in range(numdos):
    ws.cell(2 + i, 1).value = energy[i]
    if len_check == 3:
        ws.cell(2 + i, 2).value = dos[i]
        ws.cell(2 + i, 3).value = idos[i]
    else:
        ws.cell(2 + i, 2).value = dos_up[i]
        ws.cell(2 + i, 3).value = dos_down[i]
        ws.cell(2 + i, 4).value = idos_up[i]
        ws.cell(2 + i, 5).value = idos_down[i]

for i in range(int(type)):
    exec("ws.cell(1, len_check + 1 + (dos_type + 1) * i).value = \"atom_type{}\"".format(str(i + 1)))
    for j in range(dos_type):
        exec("ws.cell(1, len_check + 2 + j + (dos_type + 1) * i).value = \"DOS_type_{}-{}\"".format(str(i + 1), str(j + 1)))
        for k in range(numdos):
            exec("type{}[k][j] = float(type{}[k][j])".format(str(i + 1), str(i + 1)))
            exec("ws.cell(2 + k, len_check + 2 + j + (dos_type + 1) * i).value = type{}[k][j]".format(str(i + 1)))

chart = ScatterChart()
chart.height = 20
chart.width = 30
min_row = 2
max_row = min_row + numdos - 1
x_values = Reference(ws, min_col = 1, min_row = min_row, max_row = max_row)

for i in range(int((len_check - 1) / 2)):
    y_values = Reference(ws, min_col = 2 + i, min_row = 1, max_row = max_row)
    series = Series(y_values, x_values, title_from_data = True)
    chart.series.append(series)

for i in range(int(type)):
    for j in range(dos_type):
        y_values = Reference(ws, min_col = len_check + 2 + j + (dos_type + 1) * i, min_row = 1, max_row = max_row)
        series = Series(y_values, x_values, title_from_data = True)
        chart.series.append(series)

chart.x_axis.title = 'Energy'
chart.y_axis.title = 'DOS'
chart.x_axis.crosses = "min"
chart.y_axis.crosses = "min"
ws.add_chart(chart, "D5")

wb.save("DOS.xlsx")
